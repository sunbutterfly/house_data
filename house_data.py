
import numpy as np
import openpyxl
xls_path = '/Users/laura/desktop/2017-08_Median_Prices_of_Existing_Detached_Homes_(Historical_Data).xlsx'

def get_row_values(worksheet, row_num):
    rows = list(worksheet.iter_rows(min_row=row_num,max_row=row_num))
    return[cell.value for cell in rows[0]]

def data_for_region(xls_path, region):
    workbook = openpyxl.load_workbook(xls_path)
    sheet = workbook.get_sheet_by_name('Median Price')
    headings = get_row_values(sheet, 8)

    position = headings.index(region)

    data = {}

    for row in sheet.iter_rows(min_row=285):
        date = row[0].value
        amount = row[position].value
        if date == None:
            continue
        if amount == 'NA':
            continue
        data[date] = amount
    return data


workbook = openpyxl.load_workbook(xls_path)
sheet = workbook.get_sheet_by_name('Median Price')

def heading_values_for_user(worksheet, row_num):
    rows = list(worksheet.iter_rows(min_row=row_num,min_col=2,max_col=56))
    return[cell.value for cell in rows[0]]


headings = heading_values_for_user(sheet, 8)
headings.sort()

print('Region List:' + '\n' + '\n'.join(headings))


user_region_input = input('Please enter a region you would like to know from the list:')
user_region = str(user_region_input)
if user_region_input not in headings:
   print(user_region+ ' ' +'is not in the list')
   user_region_input = input('Please enter a region you would like to know from the list:')

user_region = str(user_region_input)
user_prediction_month_input = input('Please enter a month:')
user_prediction_year_input = input('Please enter a future or current year:')
user_prediction_month = int(user_prediction_month_input)
user_prediction_year = int(user_prediction_year_input)
user_prediction = user_prediction_month + user_prediction_year


housedata = data_for_region(xls_path,user_region)

for date, amount in sorted(housedata.items()):
   print('{0:%m}/{0:%Y}: ${1:.2f}'.format(date,amount))

x_raw_val = list(housedata.keys())
y_raw_val = list(housedata.values())

x_train_val = [
    d.month/12 + d.year
    for d in x_raw_val
]

x_train = np.array([x_train_val]).reshape(-1,1)
y_train = y_raw_val



import matplotlib.pyplot as plt
from sklearn.linear_model import LinearRegression

model = LinearRegression()
model.fit(x_train,y_train)

test_housing_data = np.array([[user_prediction]])
predited_price = model.predict(test_housing_data)[0]
print('The predicted median housing price for {0} should be ${1:.2f}'.format(user_region,predited_price) )

def f(x):
    m, b = model.coef_[0], model.intercept_
    return m * x + b

fitted_x = x_train_val
fitted_y = [f(x) for x in fitted_x]


plt.figure()
plt.plot(x_train,y_train,'k.')
plt.plot(fitted_x,fitted_y,'r')
plt.title('House Data By Region')
plt.xlabel('Date')
plt.ylabel('Amount')
plt.show()

